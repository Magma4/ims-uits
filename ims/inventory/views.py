from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from .models import *
from django.contrib.auth.decorators import login_required
from .forms import *
from django.contrib.auth.models import User
from django.db.models import Q, Sum
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import Workbook
from openpyxl.styles import *
import decimal
from datetime import datetime
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from django.db.models.functions import TruncMonth
from django.db.models import Count

from django.template.loader import render_to_string
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import os
from django.db.models.functions import ExtractMonth, ExtractDay, ExtractYear
from django.urls import reverse_lazy, reverse
from django.views import generic
from bootstrap_modal_forms.generic import (
  BSModalCreateView,
  BSModalUpdateView,
  BSModalDeleteView
)

class Index(generic.ListView):
    model = Order
    context_object_name = 'order'
    template_name = 'dashboard.html'

# Create
class OrderCreateView(BSModalCreateView):
    template_name = 'dashboard/add_request.html'
    form_class = OrderForm
    success_message = 'Order was created.'
    success_url = reverse_lazy('dashboard')

    def form_valid(self, form):
        instance = form.save(commit=False)
        instance.users = self.request.user  # Saving the current user to the order
        order_quantity = instance.order_quantity
        stock_quantity = instance.item_name.quantity

        if order_quantity <= stock_quantity:
            messages.success(self.request, "Order successfully created")
        else:
            messages.error(self.request, "Order quantity cannot be more than stock quantity")
            # Redirecting to the same page if the form is not valid
            return HttpResponseRedirect(self.request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

        return super().form_valid(form)  # This saves the form and redirects to success_url

    def get_context_data(self, **kwargs):
        context = super(OrderCreateView, self).get_context_data(**kwargs)
        context['orders'] = Order.objects.all()  # Adding orders to the context
        return context

# Update
class OrderUpdateView(BSModalUpdateView):
    model = Order
    template_name = 'dashboard/order_update.html'
    form_class = OrderForm
    success_message = 'Order was updated.'

    def form_valid(self, form):
        instance = form.save(commit=False) # Saving the current user to the order
        order_quantity = instance.order_quantity
        stock_quantity = instance.item_name.quantity

        if order_quantity <= stock_quantity:
            messages.success(self.request, "Order updated successfully")
        else:
            messages.error(self.request, "Order quantity cannot be more than stock quantity")
            # Redirecting to the same page if the form is not valid
            return HttpResponseRedirect(self.request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))

        return super().form_valid(form)  # This saves the form and redirects to success_url

    def get_context_data(self, **kwargs):
        context = super(OrderUpdateView, self).get_context_data(**kwargs)
        context['orders'] = Order.objects.all()  # Adding orders to the context
        return context
    
    def get_success_url(self):
        # Check if the user is a superuser
        if self.request.user.is_superuser:
            return reverse('view-request')  # Redirect superusers to the dashboard
        else:
            return reverse('dashboard')


# Delete
class OrderDeleteView(BSModalDeleteView):
    model = Order
    template_name = 'dashboard/order_delete.html'
    success_message = 'Order was deleted.'
    
    def get_success_url(self):
        # Check if the user is a superuser
        if self.request.user.is_superuser:
            return reverse('view-request')  # Redirect superusers to the dashboard
        else:
            return reverse('dashboard')
# Create your views here.

class StockIndex(generic.ListView):
    model = Stock
    context_object_name = 'stocks'  # It's good to use the plural form here to reflect multiple items
    template_name = 'view_stock.html'
    
class StockCreateView(BSModalCreateView):
    template_name = 'dashboard/add_stock.html'  # Ensure you create this template
    form_class = StockForm
    success_message = 'Stock was created.'
    success_url = reverse_lazy('dashboard')

    def form_valid(self, form):
        # Check if a stock with the same name (or other unique identifier) already exists
        name = form.cleaned_data.get('name')  # Assuming 'name' is a field in StockForm
        if Stock.objects.filter(name=name).exists():
            # Add an error message to the form
            messages.error(self.request, 'Product with this name already exists.')
            return self.form_invalid(form)  # Return the form with errors

        return super().form_valid(form)
    
    
    
class StockUpdateView(BSModalUpdateView):
    model = Stock
    template_name = 'dashboard/stock_update.html'
    form_class = StockForm
    success_message = 'Product was updated.'
    success_url = reverse_lazy('view-stock')

class StockDeleteView(BSModalDeleteView):
    model = Stock
    template_name = 'dashboard/stock_delete.html'
    success_message = 'Product was deleted.'
    success_url = reverse_lazy('view-stock')



@login_required(login_url='user-login')
def dashboard(request):
    user = request.user
    items = Stock.objects.all()
    order_count = Order.objects.filter(users=user).count()
    workers = User.objects.all()
    all_orders = Order.objects.all().order_by('-date')
    total_released_quantity = Order.objects.filter(status='released').aggregate(total_quantity=Sum('order_quantity'))['total_quantity'] or 0
    released_orders = Order.objects.filter(users=user, status='released').count()
    pending_orders = Order.objects.filter(users=user, status='pending').count()
    
    # Calculate percentage of released orders and pending orders
    if order_count > 0:
        released_percentage = (released_orders / order_count) * 100
        pending_percentage = (pending_orders / order_count) * 100
    else:
        released_percentage = 0
        pending_percentage = 0

    current_year = timezone.now().year
    orders_by_month = Order.objects.filter(date__year=current_year).annotate(
        month=TruncMonth('date')
    ).values('month').annotate(count=Count('id')).order_by('month')

    months = [datetime(2000, i + 1, 1).strftime('%b') for i in range(12)]  # Generate month names
    order_counts = [0] * 12  # Initialize list to hold counts for each month

    for order in orders_by_month:
        month_index = order['month'].month - 1  # Get month index (0 for Jan, 1 for Feb, etc.)
        order_counts[month_index] = order['count']

    released_items = Order.objects.filter(status='released').select_related('item_name')

    context = {
        'order_count': order_count,
        'items' : items,
        'workers' : workers,
        'months': months,
        'order_counts': order_counts,
        'all_orders': all_orders,
        'total_released_quantity': total_released_quantity,
        'released_items': released_items,
        'released_orders' : released_orders,
        'pending_orders' : pending_orders,
        'released_percentage': released_percentage,
        'pending_percentage': pending_percentage
    }
    
    return render(request, 'dashboard/dashboard.html', context)

@login_required
def viewstock(request):
    items = Stock.objects.all()
    orders = Order.objects.all()


    if request.method == 'POST':
        form = StockForm(request.POST)
        if form.is_valid():
            form.save()
            stock_name = form.cleaned_data.get('name')
            stock_quantity = form.cleaned_data.get('quantity')
            messages.success(request, f'{stock_quantity} {stock_name} has been added.')
            return redirect('stock')
    else:
        form = StockForm()
    mydictionary = {
        "stocks" : items,
        "form" : form,
        "orders" : orders
    }
    return render(request, 'dashboard/view_stock.html', context=mydictionary)

@login_required
def addstock(request):
    items = Stock.objects.all()
    orders = Order.objects.all()


    if request.method == 'POST':
        form = StockForm(request.POST)
        if form.is_valid():
            form.save()
            stock_name = form.cleaned_data.get('name')
            messages.success(request, f'{stock_name} has been added.')
            return redirect('view-stock')
    else:
        form = StockForm()
    mydictionary = {
        "stocks" : items,
        "form" : form,
        "orders" : orders
    }
    return render(request, 'dashboard/add_stock.html', context=mydictionary)


@login_required
def viewrequest(request):
    orders = Order.objects.annotate(month=ExtractMonth('date'),day=ExtractDay('date'),year=ExtractYear('date')).order_by('-month', '-day','year')
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.users = request.user
            order_quantity = instance.order_quantity
            stock_quantity = instance.item_name.quantity
            if order_quantity <= stock_quantity:
                instance.save()
                messages.success(request, "Order succesfully created")
                return redirect('dashboard')
            else:
                messages.error(request, "Order quantity cannot be more than stock quantity")
                return redirect('add-request')  # Redirect back to the requisition page
    else:
        form = OrderForm()
    context = {
        'orders': orders,
        'form': form,
    }
    return render(request, 'dashboard/view_request.html', context)


def addrequest(request):
    orders = Order.objects.all()
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.users = request.user
            order_quantity = instance.order_quantity
            stock_quantity = instance.item_name.quantity
            if instance.order_quantity <= instance.item_name.quantity:
                instance.save()
                messages.success(request, "Order successfully created")
            else:
                messages.error(request, "Order quantity cannot be more than stock quantity")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER', 'redirect_if_referer_not_found'))
    else:
        form = OrderForm()
        orders = Order.objects.all()  # If needed elsewhere

    context = {
                'form': form,
               'orders': orders
            }
    return render(request, 'dashboard/add_request_partial.html', context)


@login_required
def employees(request):
    workers = User.objects.filter(is_active=True)
    
    context = {
        'workers': workers
    }
    return render(request, 'dashboard/employees.html', context)

def instructions(request):

    return render(request, 'dashboard/instructions.html')


@login_required
def stock_delete(request, pk):
    item = Stock.objects.get(id=pk)
    if request.method == 'POST':
        item.delete()
        messages.success(request, "Item has been deleted")
        return redirect('view-stock')
    return render(request, 'dashboard/view_stock.html')

@login_required
def stock_update(request, pk):
    item = Stock.objects.get(id=pk)
    if request.method == 'POST':
        form = StockForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Stock Updated")
            return redirect('view-stock')
    else:
        form = StockForm(instance=item)
    context = {
        'form' : form
    }
    return render(request, 'dashboard/stock_update.html', context)

@login_required
def employees_detail(request, pk):
    workers = User.objects.get(id=pk)
    context={
        'workers' : workers,
    }

    return render(request, 'dashboard/employees_detail.html', context)

@login_required
def update_order_status(request, order_id):
    if request.method == 'POST' and request.user.is_superuser:
        order = Order.objects.get(id=order_id)
        new_status = request.POST.get('status')
        
        # Check if the status is being changed to 'released'
        if new_status == 'released':
            order.status = new_status
            order.released_by = request.user.username  # Set the released_by field to the username of the admin
            order.save()
            messages.success(request, f"Order {order.id} has been released by {request.user.username}.")
        elif new_status == 'returned':
            order.status = new_status
            order.returned_to = request.user.username  # Set the returned_to field to the username of the admin
            order.save()
            messages.success(request, f"Order {order.id} has been marked as received by {request.user.username}.")
        else:
            order.status = new_status
            order.save()
            messages.success(request, f"Order {order.id} status has been updated.")
    
    return redirect('view-request')


@login_required
def delete_order(request, pk):
    order = Order.objects.get(id=pk)
    if request.method == 'POST':       
        order.delete()
        if request.user.is_superuser:
            messages.success(request, f"Order has been deletd")
            return redirect('view-request')
            
        else:
            messages.success(request, f"Order has been deleted")
            return redirect('dashboard')
            
        
    return render(request, 'dashboard/view_request.html')

@login_required
def order_update(request, pk):
    order = Order.objects.get(id=pk)
    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            instance = form.save(commit=False)
            current_stock_quantity = instance.item_name.quantity
            if instance.order_quantity > current_stock_quantity:
                # Add an error message if the condition is met
                messages.error(request, "Insufficient stock quantity")
            else:
                instance.save()
                messages.success(request, f"Order {order.id} has been updated")
                if request.user.is_superuser:
                    return redirect('view-request')
                else:
                    return redirect('dashboard')
    else:
        form = OrderForm(instance=order)
    context = {
        'form': form,
    }
    return render(request, 'dashboard/order_update.html', context)

# @login_required
# def list_requisition(request):
#     return render(request, 'dashboard/list_requisition.html')

@login_required
def searchdata(request):
    q = request.GET.get('query') # Get the query parameter from the request
    if q:
        orders = Order.objects.filter(Q(users__username__icontains=q) | Q(order_description__icontains=q) | Q(users__first_name__icontains=q) | Q(users__last_name__icontains=q) | Q(id=q))
    else:
        orders = Order.objects.all()
        messages.error(request, "No results found")

    context = {
        "orders": orders,
    }
    return render(request, 'dashboard/view_request.html', context=context)

@login_required
def searchdata2(request):
    q = request.GET.get('query') # Get the query parameter from the request
    if q:
        workers = User.objects.filter(Q(username__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q))
    else:
        workers = User.objects.all()
        messages.error(request, "No results found")

    context = {
        "workers": workers,
    }
    return render(request, 'dashboard/employees.html', context=context)

@login_required
def searchdata3(request):
    q = request.GET.get('query')  # Get the query parameter from the request
    if q:
        stocks = Stock.objects.filter(Q(name__icontains=q) | Q(description__icontains=q) | Q(id=q))
    else:
        stocks = Stock.objects.all()
        messages.error(request, "No results found")

    context = {
        "stocks": stocks,
    }
    return render(request, 'dashboard/view_stock.html', context=context)

def is_valid_queryparam(param):
    return param != '' and param is not None

@login_required
def report(request):
    orders = Order.objects.annotate(
        month_year=TruncMonth('date')
    ).values('month_year').annotate(
        count=Count('id')
    ).order_by('-month_year')
    status_options = Order.objects.values_list('status', flat=True).distinct()

    available_months = [(order['month_year'].strftime('%B %Y'), order['month_year'].strftime('%Y-%m-01'), order['month_year'].strftime('%Y-%m-31')) for order in orders]

    ol = Order.objects.order_by('-id')
    orders1 = Order.objects.all()

    released_by_set = set()
    received_by_set = set()

    for order in orders1:
        if order.released_by:
            released_by_set.add(order.released_by)
        if order.returned_to:
            received_by_set.add(order.returned_to)
    
    order_id = request.GET.get('id')
    name = request.GET.get('name')
    itemName = request.GET.get('item_name')
    quantity = request.GET.get('order_quantity')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status = request.GET.get('status')
    released_by = request.GET.get('released_by')
    received_by = request.GET.get('received_by')

    request.session['id'] = order_id
    request.session['name'] = name
    request.session['item_name'] = itemName
    request.session['date_from'] = date_from
    request.session['date_to'] = date_to
    request.session['status'] = status
    request.session['released_by'] = released_by
    request.session['received_by'] = received_by

    # Filter queryset based on search parameters
    ol = ol.filter(
        Q(date__range=[date_from, date_to]) if date_from and date_to else Q()
    )
    if is_valid_queryparam(order_id):
        ol = ol.filter(id=order_id)
    if is_valid_queryparam(name):
        ol = ol.filter(users__username__icontains=name)
    if is_valid_queryparam(itemName):
        ol = ol.filter(item_name__name__icontains=itemName)
    if is_valid_queryparam(quantity):
        ol = ol.filter(order_quantity=quantity)
    if is_valid_queryparam(status):
        ol = ol.filter(status=status)
    if is_valid_queryparam(released_by):
        ol = ol.filter(released_by__icontains=released_by)
    if is_valid_queryparam(received_by):
        ol = ol.filter(returned_to__icontains=received_by)

    page = request.GET.get('page', 1)
    paginator = Paginator(ol, 30)

    try:
        ol = paginator.page(page)
    except PageNotAnInteger:
        ol = paginator.page(1)
    except EmptyPage:
        ol = paginator.page(paginator.num_pages)

    context = {
        'orders1' : orders,
        'order_list': ol,
        'order_id' : order_id,
        'name': name,
        'itemName': itemName,
        'quantity': quantity,
        'date_from': date_from,
        'date_to': date_to,
        'status': status,
        'released_by': released_by,
        'received_by': received_by,
        'available_months': available_months,
        'released_by_options': released_by_set,
        'received_by_options': received_by_set,
        'status_options' : status_options
    }
    return render(request, 'dashboard/report.html', context)

@login_required
def order_excel(request):
    ol = Order.objects.order_by('users')
    order_id = request.session.get('id')
    name = request.session.get('name')
    itemName = request.session.get('item_name')
    date_created = request.session.get('date')
    date_returned = request.session.get('returned_date')
    status = request.session.get('status')
    released_by = request.session.get('released_by')
    received_by = request.session.get('returned_to')

    if is_valid_queryparam(order_id):
        ol = ol.filter(id=order_id)
    if is_valid_queryparam(name):
        ol = ol.filter(users__username__icontains=name)
    if is_valid_queryparam(itemName):
        ol = ol.filter(item_name__name__icontains=itemName)
    if is_valid_queryparam(date_created):
        ol = ol.filter(date__icontains=date_created)
    if is_valid_queryparam(date_returned):
        ol = ol.filter(returned_date__icontains=date_returned)
    if is_valid_queryparam(status):
        ol = ol.filter(status=status)
    if is_valid_queryparam(released_by):
        ol = ol.filter(released_by__icontains=released_by)
    if is_valid_queryparam(received_by):
        ol = ol.filter(returned_to__icontains=received_by)

    order_id = order_id if order_id else "All Order ID's"
    name = name if name else "All Orders"
    itemName = itemName if itemName else "All Items"
    date_created = date_created if date_created else "2024 - 2090"
    date_returned = date_returned if date_returned else "2024 - 2090"
    status = status if status else "All Status"
    released_by = released_by if released_by else "All Admins"
    received_by = received_by if received_by else "All Admins"

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Order Report {date_from} to {date_to}.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:I1')
    worksheet.merge_cells('A2:I2')
    first_cell = worksheet.cell(row=1, column=1)
    first_cell.value = f"Report For Orders Generated on {timezone.now()}"

    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = f'Order List {date_from} to {date_to}'

    # Define the titles for columns
    columns = ['Username', 'Order ID', 'Item Name', 'Quantity', 'Date Created', 'Date Received', 'Status', 'Released By', 'Received By']
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column width
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = max(len(str(column_title)), 12)  # Set minimum width

    for order in ol:
        row_num += 1

        # Define the data for each cell in the row
        row = [order.users.username, order.id , order.item_name.name, order.order_quantity,
               order.date.replace(tzinfo=None) if order.date else None,
               order.returned_date.replace(tzinfo=None) if order.returned_date else None,
               order.status, order.released_by, order.returned_to]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if isinstance(cell_value, datetime):
                cell.number_format = 'yyyy-mm-dd HH:MM:SS'  # Set datetime format

            # Adjust column width
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(worksheet.column_dimensions[column_letter].width, len(str(cell_value)) + 2)  # Set minimum width

        # Adjust row height based on content
        worksheet.row_dimensions[row_num].height = 14.4  

    # Write total count to the last cell
    total_count_cell = worksheet.cell(row=row_num + 1, column=1)
    total_count_cell.value = "Total Count"
    total_count_cell.font = Font(bold=True)
    total_count_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    total_count_value_cell = worksheet.cell(row=row_num + 1, column=2)
    total_count_value_cell.value = len(ol)
    total_count_value_cell.font = Font(bold=True)
    total_count_value_cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(response)
    return response

@login_required
def order_pdf(request):
    # Your existing code to fetch orders
    ol = Order.objects.order_by('id')
    # Retrieve filters from session
    order_id = request.session.get('id')
    name = request.session.get('name')
    itemName = request.session.get('item_name')
    date_created = request.session.get('date')
    date_returned = request.session.get('returned_date')
    status = request.session.get('status')
    released_by = request.session.get('released_by')
    received_by = request.session.get('returned_to')
    # Apply filters to queryset
    if is_valid_queryparam(order_id):
        ol = ol.filter(id=order_id)
    if is_valid_queryparam(name):
        ol = ol.filter(users__username__icontains=name)
    if is_valid_queryparam(itemName):
        ol = ol.filter(item_name__name__icontains=itemName)
    if is_valid_queryparam(date_created):
        ol = ol.filter(date__icontains=date_created)
    if is_valid_queryparam(date_returned):
        ol = ol.filter(returned_date__icontains=date_returned)
    if is_valid_queryparam(status):
        ol = ol.filter(status=status)
    if is_valid_queryparam(released_by):
        ol = ol.filter(released_by__icontains=released_by)
    if is_valid_queryparam(received_by):
        ol = ol.filter(returned_to__icontains=received_by)

    # Create a PDF report
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Order_Report.pdf"'

    # Define the data for PDF
    data = []
    header = ['Username', 'Order ID', 'Item Name', 'Quantity', 'Date Created', 'Date Received', 'Status', 'Released By', 'Received By']
    data.append(header)
    for order in ol:
        data.append([order.users.username, order.id, order.item_name.name, order.order_quantity,
                     order.date.replace(tzinfo=None) if order.date else None,
                     order.returned_date.replace(tzinfo=None) if order.returned_date else None,
                     order.status, order.released_by, order.returned_to])

    # Calculate total count of orders
    total_count = len(ol)

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=landscape(letter), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    
    table_style = TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])

    table = Table(data, style=table_style)
    styles = getSampleStyleSheet()
    
    elements = []
    elements.append(Paragraph(f"A Report of Orders Generated on {timezone.now()}", styles['title']))
    elements.append(Paragraph(f"Filters used:", styles['title']))
    elements.append(Paragraph(f"Order ID: {order_id if order_id else 'All'}, Name: {name if name else 'All'}, Item Name: {itemName if itemName else 'All'}, Date Created: {date_created if date_created else 'All'}, Date Returned: {date_returned if date_returned else 'All'}, Status: {status if status else 'All'}, Released By: {released_by if released_by else 'All'}, Received By: {received_by if received_by else 'All'}", styles['Normal']))
    elements.append(table)
    elements.append(Paragraph(f"Total Count of Orders: {total_count}", styles['Normal']))
    
    doc.build(elements)
    return response